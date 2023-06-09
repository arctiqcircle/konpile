#!/usr/bin/env python3

import sys
import os
from pathlib import Path
import argparse
from datetime import datetime

import jinja2
import openpyxl
import csv
import json

DataSet = dict[str, list[dict]]


def std_str(obj: any):
    obj = str(obj)
    if obj is not None:
        return (
            obj.lower()
            .replace("-", "_")
            .replace(" ", "_")
            .replace(",", "_")
            .replace(".", "_")
        )


def process_file(datafile: Path) -> DataSet:
    """
    Determine the type of the file and the appropriate processing function
    using a match statement, then return the processed data.
    """
    # TODO: Implement file parsing injection for decoupling.
    match datafile.suffix.lower():
        case ".xlsx":
            return process_excel(datafile)
        case ".csv":
            return process_csv(datafile)
        case ".json":
            return process_json(datafile)
        case _:
            raise ValueError(f"Unsupported file type: {datafile.suffix}")


def process_excel(datafile: Path) -> DataSet:
    """
    Unpack an Excel file into a dictionary of lists of dictionaries (DataSet).
    The schema is SHEETNAME: [ {COLUMNNAME: VALUE, ...}, ... ]
    """
    # Load workbook
    workbook = openpyxl.load_workbook(datafile)
    # Load sheets
    sheets = [workbook[sheet] for sheet in workbook.sheetnames]
    # Build dataset
    dataset = {}
    for sheet in sheets:
        headers = [std_str(cell.value) for cell in sheet[1]]
        dataset[std_str(sheet.title)] = [
            {
                headers[i]: cell.value
                for i, cell in enumerate(row)
            }
            for row in sheet.iter_rows(min_row=2)
        ]
    return dataset


def process_csv(datafile: Path) -> DataSet:
    """
    Unpack a CSV file into a dictionary of lists of dictionaries (DataSet).
    The schema is FILENAME: [ {COLUMNNAME: VALUE, ...}, ... ]
    """
    # Load CSV file
    with open(datafile, "r") as f:
        reader = csv.DictReader(f)
        dataset = {std_str(datafile.stem): [row for row in reader]}
    return dataset


def process_json(datafile: Path) -> DataSet:
    """
    Unpack a JSON file into a dictionary of lists of dictionaries (DataSet).
    The schema is FILENAME: [ {COLUMNNAME: VALUE, ...}, ... ]
    """
    # Load JSON file
    with open(datafile, "r") as f:
        data = json.load(f)
        dataset = {std_str(datafile.stem): data}
    return dataset


def render_file(template_file: Path, data: DataSet):
    """
    Render a Jinja2 template with data.
    """
    env = jinja2.Environment(
        loader=jinja2.FileSystemLoader(searchpath=str(template_file.parent))
    )
    template = env.get_template(template_file.name)
    rendering = template.render(data)
    return rendering


def render_string(template_string: str, data: DataSet):
    """
    Render a Jinja2 template string with data.
    """
    env = jinja2.Environment(loader=jinja2.BaseLoader())
    template = env.from_string(template_string)
    rendering = template.render(data)
    return rendering


def translate_fields(dataset: any, translators: list[Path]):
    """
    Translate fields in the dataset using a list of CSV files.
    The CSV files must have two columns: original and translation.
    This allows us to restructure data without having to affect
    the original files. If we find a field that matches the
    name of the file (without the extension), we translate the
    values in that field using the CSV file.
    """
    for translator in translators:
        with open(translator, "r") as f:
            target_field = std_str(translator.stem)
            reader = csv.DictReader(f)
            translation_table = {row["original"]: row["translation"] for row in reader}
            def translate(field_name: str, obj: any):
                # Recursively translate all applicable values in an object
                # whenever the key matches the name of the translator file.
                if isinstance(obj, dict):
                    return {
                        key: translate(key, value)
                        for key, value in obj.items()
                    }
                elif isinstance(obj, list):
                    return [translate(field_name, item) for item in obj]
                else:
                    return translation_table[str(obj)] if field_name == target_field else obj
            dataset = translate(None, dataset)
    return dataset


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Use an Excel file to fill a Jinja2 template."
    )
    parser.add_argument("template_file", type=Path, help="Jinja2 template file")
    parser.add_argument("data_file", type=Path, help="File with data")
    parser.add_argument(
        "--translation", "-t", type=Path, nargs="+", help="Translation CSV files"
    )
    args = parser.parse_args(sys.argv[1:])
    print("Unpacking data...")
    data = process_file(args.data_file)
    if args.translation:
        print("Translating fields...")
        data = translate_fields(data, args.translation)
    print("Rendering template...")
    rendering = render_file(args.template_file, data)
    # print(rendering)
    with open(
        f'{args.template_file.stem}_{datetime.now().strftime("%Y%m%d%H%M%S")}.txt', "w"
    ) as output_file:
        output_file.write(rendering)
